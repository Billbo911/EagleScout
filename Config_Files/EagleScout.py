import os
import glob
import pandas as pd
import sqlite3
import openpyxl
from openpyxl.styles import Color, PatternFill #, Font, Border
#from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import numpy as np


os.chdir ("/EagleScout")
path = '.'

#------------------------------- Perform file Clean up -------------------------------

# Remove the combined .csv file from previous runs
#This will provide clean date without corruption from earlier runs
if os.path.exists('./Spreadsheets/combined.csv'): 
    os.remove ('./Spreadsheets/combined.csv')

#Remove previous copy of the Database version
if os.path.exists('./DataBases/Combined_Raw.db'): 
    os.remove ('./DataBases/Combined_Raw.db')
    
#Remove previous Excel spreadsheet
if os.path.exists('./Spreadsheets/Tournament.xlsx'): 
    os.remove ('./Spreadsheets/Tournament.xlsx')
    
#Remove sorted combined spreadsheet
if os.path.exists('./Spreadsheets/Combined.xlsx'): 
    os.remove ('./Spreadsheets/Combined.xlsx')
    
#Remove previous Excel spreadsheet
if os.path.exists('./Spreadsheets/Master.xlsx'): 
    os.remove ('./Spreadsheets/Master.xlsx')
        
#Remove previous Excel spreadsheet
if os.path.exists('./Spreadsheets/Temp.xlsx'): 
    os.remove ('./Spreadsheets/Temp.xlsx')
 
#Remove old copy of Match Schedule .csv    
if os.path.exists('./Spreadsheets/Match_Schdule.csv'): 
    os.remove ('./Spreadsheets/Match_Schdule.csv')
#----------------------------- End of file Clean up ------------------------------------

#Create the Database file
con = sqlite3.connect('./DataBases/Combined_Raw.db')



#------------------------------Conditional Formatting values----------------------------

# Create fill
redFill = PatternFill(start_color='F95555', end_color='F95555', fill_type='solid')
yellowFill = PatternFill(start_color='FBFE46', end_color='FBFE46', fill_type='solid')
greenFill = PatternFill(start_color='ACF99D', end_color='ACF99D', fill_type='solid')
clearFill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')


######################################################################################

#--------------------------------Start Building the desired files----------------------
#Read in and merge all .CSV file names
files_in_dir = [ f for f in glob.glob('*.csv')] 

StrToInt_dict = {'Team': int,'Match_Num':int,'Start_lvl': int,'A_R_C':int,'A_R_C_F':int,
'A_R_H':int,'A_R_H_F':int,'A_C_C':int,'A_C_C_F':int,'A_C_H':int,'A_C_H_F':int,
'T_R_C':int,'T_R_C_F':int,'T_R_H':int,'T_R_H_F':int,'T_C_C':int,'T_C_C_F':int,
'T_C_H':int,'T_C_H_F':int,'Drvr_Perf':int,'Auto_Perf':int,'Climb_Lvl':int,
'Point_Contrib':int,'Name':str,'Comments':str}

#StrToInt_dict = {'Team': int,'Match_Num':str,'Start_lvl': str,'A_R_C':str,'A_R_C_F':str,
#'A_R_H':str,'A_R_H_F':str,'A_C_C':str,'A_C_C_F':str,'A_C_H':str,'A_C_H_F':str,
#'T_R_C':str,'T_R_C_F':str,'T_R_H':str,'T_R_H_F':str,'T_C_C':str,'T_C_C_F':str,
#'T_C_H':str,'T_C_H_F':str,'Drvr_Perf':str,'Auto_Perf':str,'Climb_Lvl':str,
#'Point_Contrib':str,'Name':str,'Comments':str}
    

#Create a single combined .csv file with all data
#from all matches completed so far.
#and add column headers as labels
d1 = pd.read_csv('./Config_Files/Header.txt')
d1.to_csv('./Spreadsheets/combined.csv', header = True, index = False)


for filenames in files_in_dir: 
    df = pd.read_csv(filenames)
    fName, fExt = (os.path.splitext(filenames))
    sName = fName.split('-')
    N=(sName[1])
    df.insert(0,N,N,True)
    df.to_csv('./Spreadsheets/combined.csv', index_label = (sName[0]), mode = 'a')



#------------------------------------------------------------------------------------------

#Convert combined Raw csv file into one master Raw Excel Data file
#Add score contribution to each entry
with pd.ExcelWriter('./Spreadsheets/Combined.xlsx', engine = 'xlsxwriter') as writer:
    dt = pd.read_csv('./Spreadsheets/combined.csv')
    rows = len(dt.index)
    
    dt.to_excel(writer, sheet_name = 'All data',index = False)
    worksheet = writer.sheets['All data']
    #--------------------Add custom calculated data--------------------------
    for i in range(1,rows+1):   #Add Score contribution
        SC = (dt.at[-1+i,'A_R_H'] *2)
        SC += (dt.at[-1+i,'A_R_C'] *3)
        SC += (dt.at[-1+i,'T_R_H'] *2)
        SC += (dt.at[-1+i,'T_R_C'] *3)
        SC += (dt.at[-1+i,'A_C_H'] *2)
        SC += (dt.at[-1+i,'A_C_C'] *3)
        SC += (dt.at[-1+i,'T_C_H'] *2)
        SC += (dt.at[-1+i,'T_C_C'] *3)
        worksheet.write( i , 24 , SC)
    
    writer.save()
    
#------------------------------------Save Raw data to Database file-----------------------
#Save Combined.xlsx to a Database file
db = pd.read_excel('./Spreadsheets/Combined.xlsx')
db.to_sql("Raw_Data", con, if_exists = 'replace', index = False)


#Parse through Combined.xlsx files and append content to appropriate team worksheet.
#Read in the file and set the values to 'int'
with pd.ExcelWriter('./Spreadsheets/Master.xlsx') as writer:
       
    df2 = pd.read_excel('./Spreadsheets/Combined.xlsx', converters = StrToInt_dict)
    group = df2.groupby('Team')
    for Team, Team_df in group:
        
        Team_df.to_excel(writer, sheet_name = ("T"+str(Team)),index = False)
        
        
    writer.save()



#----------------------------Add formulas and manipulations ------------------------------



#Add formulas to each sheet for calculating values
Data = pd.ExcelFile('./Spreadsheets/Master.xlsx')
Teams = Data.sheet_names

Tnmt = openpyxl.load_workbook('./Spreadsheets/Master.xlsx', read_only = False, keep_vba = True)

#Add new worksheet for easy access to pertinant info
WS1 = Tnmt.create_sheet("Important Stuff", 0)
WS2 = Tnmt.create_sheet("Predictions")

#Add the same formulas to each team's sheet
for sht in Teams:
    sn = Tnmt.get_sheet_by_name(sht)
    sn['B16']= str('Average')
    sn['B17']= str('Standard Deviation')
    sn['B18']= str('STDev % of Average')
    sn['B19']= str('Total')
    sn['E21']= str('Rocket Hatch Ave')
    sn['E22']= str('Cargo Hatch Ave')
    sn['E23']= str('Rocket Cargo Ave')
    sn['E24']= str('Cargo Cargo Ave')
    sn['E25']= str('Climb Level Ave')
    
    sn.cell(row = 16, column = 3).value=  "=AVERAGE(C2:C13)"
    sn.cell(row = 16, column = 4).value=  "=AVERAGE(D2:D13)"
    sn.cell(row = 16, column = 5).value=  "=AVERAGE(E2:E13)"
    sn.cell(row = 16, column = 6).value=  "=AVERAGE(F2:F13)"
    sn.cell(row = 16, column = 7).value=  "=AVERAGE(G2:G13)"
    sn.cell(row = 16, column = 8).value=  "=AVERAGE(H2:H13)"
    sn.cell(row = 16, column = 9).value=  "=AVERAGE(I2:I13)"
    sn.cell(row = 16, column = 10).value=  "=AVERAGE(J2:J13)"
    sn.cell(row = 16, column = 11).value=  "=AVERAGE(K2:K13)"
    sn.cell(row = 16, column = 12).value=  "=AVERAGE(L2:L13)"
    sn.cell(row = 16, column = 13).value=  "=AVERAGE(M2:M13)"
    sn.cell(row = 16, column = 14).value=  "=AVERAGE(N2:N13)"
    sn.cell(row = 16, column = 15).value=  "=AVERAGE(O2:O13)"
    sn.cell(row = 16, column = 16).value=  "=AVERAGE(P2:P13)"
    sn.cell(row = 16, column = 17).value=  "=AVERAGE(Q2:Q13)"
    sn.cell(row = 16, column = 18).value=  "=AVERAGE(R2:R13)"
    sn.cell(row = 16, column = 19).value=  "=AVERAGE(S2:S13)"
    sn.cell(row = 16, column = 20).value=  "=AVERAGE(T2:T13)"
    sn.cell(row = 16, column = 21).value=  "=AVERAGE(U2:U13)"
    sn.cell(row = 16, column = 22).value=  "=AVERAGE(V2:V13)"
    sn.cell(row = 16, column = 25).value=  "=AVERAGE(Y2:Y13)"
    sn.cell(row = 17, column = 3).value=  "=STDEV(C2:C13)"
    sn.cell(row = 17, column = 4).value=  "=STDEV(D2:D13)"
    sn.cell(row = 17, column = 5).value=  "=STDEV(E2:E13)"
    sn.cell(row = 17, column = 6).value=  "=STDEV(F2:F13)"
    sn.cell(row = 17, column = 7).value=  "=STDEV(G2:G13)"
    sn.cell(row = 17, column = 8).value=  "=STDEV(H2:H13)"
    sn.cell(row = 17, column = 9).value=  "=STDEV(I2:I13)"
    sn.cell(row = 17, column = 10).value=  "=STDEV(J2:J13)"
    sn.cell(row = 17, column = 11).value=  "=STDEV(K2:K13)"
    sn.cell(row = 17, column = 12).value=  "=STDEV(L2:L13)"
    sn.cell(row = 17, column = 13).value=  "=STDEV(M2:M13)"
    sn.cell(row = 17, column = 14).value=  "=STDEV(N2:N13)"
    sn.cell(row = 17, column = 15).value=  "=STDEV(O2:O13)"
    sn.cell(row = 17, column = 16).value=  "=STDEV(P2:P13)"
    sn.cell(row = 17, column = 17).value=  "=STDEV(Q2:Q13)"
    sn.cell(row = 17, column = 18).value=  "=STDEV(R2:R13)"
    sn.cell(row = 17, column = 19).value=  "=STDEV(S2:S13)"
    sn.cell(row = 17, column = 20).value=  "=STDEV(T2:T13)"
    sn.cell(row = 17, column = 21).value=  "=STDEV(U2:U13)"
    sn.cell(row = 17, column = 22).value=  "=STDEV(V2:V13)"
    sn.cell(row = 17, column = 25).value=  "=STDEV(Y2:Y13)"
    sn.cell(row = 18, column = 3).value=  "=SUM(C17/C16)"
    sn.cell(row = 18, column = 4).value=  "=SUM(D17/D16)"
    sn.cell(row = 18, column = 5).value=  "=SUM(E17/E16)"
    sn.cell(row = 18, column = 6).value=  "=SUM(F17/F16)"
    sn.cell(row = 18, column = 7).value=  "=SUM(G17/G16)"
    sn.cell(row = 18, column = 8).value=  "=SUM(H17/H16)"
    sn.cell(row = 18, column = 9).value=  "=SUM(I17/I16)"
    sn.cell(row = 18, column = 10).value=  "=SUM(J17/J16)"
    sn.cell(row = 18, column = 11).value=  "=SUM(K17/K16)"
    sn.cell(row = 18, column = 12).value=  "=SUM(L17/L16)"
    sn.cell(row = 18, column = 13).value=  "=SUM(M17/M16)"
    sn.cell(row = 18, column = 14).value=  "=SUM(N17/N16)"
    sn.cell(row = 18, column = 15).value=  "=SUM(O17/O16)"
    sn.cell(row = 18, column = 16).value=  "=SUM(P17/P16)"
    sn.cell(row = 18, column = 17).value=  "=SUM(Q17/Q16)"
    sn.cell(row = 18, column = 18).value=  "=SUM(R17/R16)"
    sn.cell(row = 18, column = 19).value=  "=SUM(S17/S16)"
    sn.cell(row = 18, column = 20).value=  "=SUM(T17/T16)"
    sn.cell(row = 18, column = 21).value=  "=SUM(U17/U16)"
    sn.cell(row = 18, column = 22).value=  "=SUM(V17/V16)"
    sn.cell(row = 19, column = 3).value=  "=SUM(C2:C13)"
    sn.cell(row = 19, column = 4).value=  "=SUM(D2:D13)"
    sn.cell(row = 19, column = 5).value=  "=SUM(E2:E13)"
    sn.cell(row = 19, column = 6).value=  "=SUM(F2:F13)"
    sn.cell(row = 19, column = 7).value=  "=SUM(G2:G13)"
    sn.cell(row = 19, column = 8).value=  "=SUM(H2:H13)"
    sn.cell(row = 19, column = 9).value=  "=SUM(I2:I13)"
    sn.cell(row = 19, column = 10).value=  "=SUM(J2:J13)"
    sn.cell(row = 19, column = 11).value=  "=SUM(K2:K13)"
    sn.cell(row = 19, column = 12).value=  "=SUM(L2:L13)"
    sn.cell(row = 19, column = 13).value=  "=SUM(M2:M13)"
    sn.cell(row = 19, column = 14).value=  "=SUM(N2:N13)"
    sn.cell(row = 19, column = 15).value=  "=SUM(O2:O13)"
    sn.cell(row = 19, column = 16).value=  "=SUM(P2:P13)"
    sn.cell(row = 19, column = 17).value=  "=SUM(Q2:Q13)"
    sn.cell(row = 19, column = 18).value=  "=SUM(R2:R13)"
    sn.cell(row = 19, column = 19).value=  "=SUM(S2:S13)"
    sn.cell(row = 19, column = 20).value=  "=SUM(T2:T13)"
    sn.cell(row = 19, column = 21).value=  "=SUM(U2:U13)"
    sn.cell(row = 19, column = 22).value=  "=SUM(V2:V13)"
    sn.cell(row = 21, column = 6).value=  "=SUM((F16+N16)/2)"
    sn.cell(row = 22, column = 6).value=  "=SUM((J16+R16)/2)"
    sn.cell(row = 23, column = 6).value=  "=SUM((D16+L16)/2)"
    sn.cell(row = 24, column = 6).value=  "=SUM((H16+P16)/2)"
    sn.cell(row = 25, column = 6).value=  "=SUM(V16*1)"
    



C = 0   
for sht in Teams:
    WS1.cell(row = 2+C , column = 1).value = (sht[1:])
    C += 1  #Add Header info to "Important Stuff" sheet    
Stuff = ['Team #', 'Av Roc Hatch', 'Av Roc Cargo', 'Av Carg Hatch', 'Av Carg Cargo', 'Av Climb', 'Av Point Contrib']
s = 0
for st in Stuff:
    WS1.cell(row = 1, column = 1 + s).value = str(Stuff[s])
    s += 1
    
Tnmt.save('./Spreadsheets/Temp.xlsx')

#Copy data from sheets and cells to where it's needed
Trnmt = openpyxl.load_workbook('./Spreadsheets/Temp.xlsx', read_only = False, keep_vba = True, data_only = False)

SN = Trnmt.get_sheet_by_name( 'Important Stuff')

D = 0
for tn in Teams:

    TT = "="+tn+"!"+"F21"
    TU = "="+tn+"!"+"F22"
    TV = "="+tn+"!"+"F23"
    TW = "="+tn+"!"+"F24"
    TX = "="+tn+"!"+"F25"
    TY = "="+tn+"!"+"Y16"
    SN.cell(row = 2+D , column = 2).value = TT #Average Rocket Hatch
    SN.cell(row = 2+D , column = 3).value = TU #Average Rocket Cargo
    SN.cell(row = 2+D , column = 4).value = TV #Average Cargo Ship Hatch
    SN.cell(row = 2+D , column = 5).value = TW #Average Cargo Ship Cargo
    SN.cell(row = 2+D , column = 6).value = TX #Average Climb level
    SN.cell(row = 2+D , column = 7).value = TY #Average Point Contribution    
    #SN.cell(row = 2+D , column = 3).number_format = '#,##0.00'
    
    D += 1

SN.conditional_formatting.add('B2:F75', CellIsRule(operator='greaterThan', formula= [2.96] , stopIfTrue=True, fill=greenFill))
SN.conditional_formatting.add('B2:F75', CellIsRule(operator='between', formula= [1.96,2.95] , stopIfTrue=True, fill=yellowFill))
SN.conditional_formatting.add('B2:F75', CellIsRule(operator='between', formula= [.0001,1.95] , stopIfTrue=True, fill=redFill))


Trnmt.save('./Spreadsheets/Tournament.xlsx')




#-----------------------Get data from other Excel files and add to Tournament.xls------------

book = openpyxl.load_workbook('./Spreadsheets/Tournament.xlsx')
writer = pd.ExcelWriter('./Spreadsheets/Tournament.xlsx', engine = 'openpyxl')
writer.book = book

df1 = pd.read_excel('/EagleScout/Python Scripts/Match_Schedule/Match_Schdule.xlsx', sheet_name = 'Match Schedule', header = 0)
df2 = pd.read_excel('/EagleScout/Python Scripts/Match_Schedule/Match_Schdule.xlsx', sheet_name = '2073 Schedule', header = 0)
df3 = pd.read_excel('./Spreadsheets/Tournament.xlsx', sheet_name = 'Predictions')
df1.to_excel(writer, sheet_name = 'Match Schedule', index = False)
df2.to_excel(writer, sheet_name = '2073 Schedule',index = False)

prd = book.get_sheet_by_name('Predictions')
prd.cell(row = 1, column = 1).value = 'Match #'
prd.cell(row = 1, column = 2).value = 'Red'
prd.cell(row = 1, column = 3).value = 'Blue'



#---------------------- Create Prdictions for our matches -------------------
rows = len(df2.index)
for r in range ( 0, rows):
    m = df2.at[r, 'Match #']
    prd.cell(row = r+2, column = 1 ).value = m
 
 #-------- Add score contributions values for each team on an alliance ------   
for m in range(0, rows):
    R1 = "=T"+str(df2.at[0+int(m), 'Red 1'])+"!L16"
    R2 = "=T"+str(df2.at[0+int(m), 'Red 2'])+"!L16"
    R3 = "=T"+str(df2.at[0+int(m), 'Red 3'])+"!L16"
    prd.cell(row = 100+(m*3), column = 5).value = R1
    prd.cell(row = 101+(m*3), column = 5).value = R2
    prd.cell(row = 102+(m*3), column = 5).value = R3
    prd.cell(row = 2+m, column = 2).value = "=SUM(E"+str(100+(m*3))+":E"+str(102+(m*3))
    B1 = "=T"+str(df2.at[0+int(m), 'Blue 1'])+"!L16"
    B2 = "=T"+str(df2.at[0+int(m), 'Blue 2'])+"!L16"
    B3 = "=T"+str(df2.at[0+int(m), 'Blue 3'])+"!L16"
    prd.cell(row = 100+(m*3), column = 6).value = B1
    prd.cell(row = 101+(m*3), column = 6).value = B2
    prd.cell(row = 102+(m*3), column = 6).value = B3
    prd.cell(row = 2+m, column = 3).value = "=SUM(F"+str(100+(m*3))+":F"+str(102+(m*3))


writer.save()
writer.close()


#----------------------------Create a Match Schedule.csv for use by Scouting Tablets-------------------------
MS = pd.read_excel('/EagleScout/Python Scripts/Match_Schedule/Match_Schdule.xlsx', sheet_name = 'Match Schedule')
MS.to_csv('./Spreadsheets/Match_Schdule.csv', columns = ('Match #', 'Red 1','Red 2', 'Red 3', 'Blue 1', 'Blue 2', 'Blue 3'),index = False)
