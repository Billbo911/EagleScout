import os
import tbapy
from openpyxl import Workbook
from openpyxl.styles import PatternFill #, Font, Border
from openpyxl.formatting.rule import CellIsRule #, FormulaRule, ColorScaleRule
import datetime


#------------------Set year, event code, Team number, and Time Zone offset here.------------------
Year = 2020  #Enter the year foe which you want the schedule
event = 'utwv'  #'cafr' #'idbo'  #Enter the event code for the tournament you want the Match Schedule

TEAM = 2122  #Enter your team number here to get a sheet with just your team's match schedule

UBToffset = -7  #Enter UBT Time zone offset for your location
#--------------------------------------------------------------------------------------------------------------


Event = str(Year)+event
#---------------------------------Clean up old files-----------------
#This will provide clean date without corruption from earlier runs
if os.path.exists('/EagleScout/Python Scripts/Match_Schedule/Match_Schdule.xlsx'):
    os.remove ('/EagleScout/Python Scripts/Match_Schedule/Match_Schdule.xlsx')
    
if os.path.exists('/EagleScout/Python Scripts/Match_Schedule/Match_Schdule.csv'):
    os.remove ('/EagleScout/Python Scripts/Match_Schedule/Match_Schdule.csv')
#---------------------------------End of Clean up--------------------- 

#--------------------------------Fill pattern for conditional formatting--------------------
greenFill = PatternFill(start_color='ACF99D', end_color='ACF99D', fill_type='solid')

#My personal TBA Key
tba = tbapy.TBA('P4FqDGYHIGbG9A64B2TXwYqbvb30FiO2HLyZzzUtcMKngfUmmzZlP9ZmutR1UGnf')

#----------Create workbook with two worksheets. One for the tournament, one for your Team---------
Tnmt = Workbook()
ws = Tnmt.active
ws.title = "Match Schedule"
ws2 = Tnmt.create_sheet( str(TEAM) + ' Schedule')

def Alpha_scrape(value):
    if (value.find('B') != -1):
        value = (value[:len(value)-1]) #Slice off the Alpha
        if (len(value) == 3): # 3 digit team
            value = ('9'+value)
        elif (len(value) == 2): # 2 digit team
            value = ('90' + value)
        elif (len(value) == 1): # 1 digit team
            value = ('900' + value)
        else: # 4 digit team
            value = ('9'+value[1:])
        #print(str(value))
        return value
    elif (value.find('C') != -1):
        value = (value[:len(value)-1])  #Slice off the Alpha
        if (len(value) == 3): # 3 digit team
            value = ('8'+value)
        elif (len(value) == 2): # 2 digit team
            value = ('80' + value)
        elif (len(value) == 1): # 1 digit team
            value = ('800' + value)
        else:
            value = ('8'+value[1:])
        #print(str(value))
        return value

    else:
        return value
        #print ("It not B Here")


#Query TBA for match schedule
schedule_keys = tba.event_matches(Event, simple=True, keys=True)
#print(schedule_keys)
Q = 0
for i in range (1, len(schedule_keys)):
    result = tba.match(schedule_keys[i], simple=True)
    if  result.comp_level == 'qm':
        Q += 1
#print(Q)
S = 1
for i in range (1, Q+1):
    results = tba.match(Event + '_qm' + str(i), simple=True)
    if results.comp_level == 'qm' :
        mn = results.match_number
        tu = results.predicted_time
        tm = datetime.datetime.utcfromtimestamp(tu +(UBToffset*(3600))) #Convert UNIX time to local time
        R1 = results.alliances['red']['team_keys'][0]
        R1 = R1[3:]
        R1 = Alpha_scrape(R1)
        R2 = results.alliances['red']['team_keys'][1]
        R2 = R2[3:]
        R2 = Alpha_scrape(R2)
        R3 = results.alliances['red']['team_keys'][2]
        R3 = R3[3:]
        R3 = Alpha_scrape(R3)
        B1 = results.alliances['blue']['team_keys'][0]
        B1 = B1[3:]
        B1 = Alpha_scrape(B1)
        B2 = results.alliances['blue']['team_keys'][1]
        B2 = B2[3:]
        B2 = Alpha_scrape(B2)
        B3 = results.alliances['blue']['team_keys'][2]
        B3 = B3[3:]
        B3 = Alpha_scrape(B3)
        ws.cell(row = 1+S, column = 1).value = int(mn)
        ws.cell(row = 1+S, column = 2).value = tm
        ws.cell(row = 1+S, column = 3).value = int(R1)
        ws.cell(row = 1+S, column = 4).value = int(R2)
        ws.cell(row = 1+S, column = 5).value = int(R3)
        ws.cell(row = 1+S, column = 6).value = int(B1)
        ws.cell(row = 1+S, column = 7).value = int(B2)
        ws.cell(row = 1+S, column = 8).value = int(B3)
        ws.cell(row = 1+S, column = 2).number_format = 'HH:MM' #Restrict to display time only
        S += 1
H = 0
Headers = ['Match #', 'Time', 'Red 1', 'Red 2', 'Red 3', 'Blue 1', 'Blue 2', 'Blue 3']
for h in Headers:
    ws.cell(row =1, column = 1 + H).value = Headers[H]
    ws2.cell(row =1, column = 1 + H).value = Headers[H]
    H += 1


S = 1
Tm_Sched = tba.team_matches(TEAM, Event, simple = True, keys = True)
#print(Tm_Sched)

for i in range (1, Q+1):
    if  Event + '_qm' + str(i) in Tm_Sched:
        #print(Event + '_qm' + str(i))
        results = tba.match(Event + '_qm' + str(i), simple=True)
        mn = results.match_number
        tu = results.predicted_time
        tm = datetime.datetime.utcfromtimestamp(tu +(UBToffset*(3600))) #Convert UNIX time to local time
        R1 = results.alliances['red']['team_keys'][0]
        R1 = R1[3:]
        R1 = Alpha_scrape(R1)
        R2 = results.alliances['red']['team_keys'][1]
        R2 = R2[3:]
        R2 = Alpha_scrape(R2)
        R3 = results.alliances['red']['team_keys'][2]
        R3 = R3[3:]
        R3 = Alpha_scrape(R3)
        B1 = results.alliances['blue']['team_keys'][0]
        B1 = B1[3:]
        B1 = Alpha_scrape(B1)
        B2 = results.alliances['blue']['team_keys'][1]
        B2 = B2[3:]
        B2 = Alpha_scrape(B2)
        B3 = results.alliances['blue']['team_keys'][2]
        B3 = B3[3:]
        B3 = Alpha_scrape(B3)
        
        ws2.cell(row = 1+S, column = 1).value = int(mn)
        ws2.cell(row = 1+S, column = 2).value = tm
        ws2.cell(row = 1+S, column = 3).value = int(R1)
        ws2.cell(row = 1+S, column = 4).value = int(R2)
        ws2.cell(row = 1+S, column = 5).value = int(R3)
        ws2.cell(row = 1+S, column = 6).value = int(B1)
        ws2.cell(row = 1+S, column = 7).value = int(B2)
        ws2.cell(row = 1+S, column = 8).value = int(B3)
        ws2.cell(row = 1+S, column = 2).number_format = 'HH:MM' #Restrict to display time only
        S += 1
        

SN = Tnmt.get_sheet_by_name( str(TEAM) + ' Schedule')
SN.conditional_formatting.add('C2:H14', CellIsRule(operator='equal', formula= [TEAM] , stopIfTrue=True, fill=greenFill))    

SM = Tnmt.get_sheet_by_name( 'Match Schedule')
SM.conditional_formatting.add('C2:H130', CellIsRule(operator='equal', formula= [TEAM] , stopIfTrue=True, fill=greenFill))


Tnmt.save('/EagleScout/Python Scripts/Match_Schedule/Match_Schdule.xlsx')


