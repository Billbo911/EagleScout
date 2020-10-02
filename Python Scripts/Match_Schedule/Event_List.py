#import os
import tbapy
from openpyxl import Workbook


YEAR = '2020'

Sched = Workbook()
ws = Sched.active
ws.title = "Scheduled Tournaments"

#My personal TBA Key
tba = tbapy.TBA('P4FqDGYHIGbG9A64B2TXwYqbvb30FiO2HLyZzzUtcMKngfUmmzZlP9ZmutR1UGnf')

#Query TBA for match schedule
schedule_keys = tba.events(YEAR, simple=False, keys=False)
for i in range (1, len(schedule_keys)):
    cd = schedule_keys[i].event_code
    nm = schedule_keys[i].name
    sd = schedule_keys[i].start_date
    tp = schedule_keys[i].event_type_string
    ws.cell(row = 1+i, column = 1).value = cd
    ws.cell(row = 1+i, column = 2).value = nm
    ws.cell(row = 1+i, column = 3).value = sd
    ws.cell(row = 1+i, column = 4).value = tp
   #print(schedule_keys[i].event_code + ' , ' +schedule_keys[i].name)
    
Sched.save('/EagleScout/Tournaments_Info/Tournamet_Codes.xlsx')
